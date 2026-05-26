from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy.orm import Session
from sqlalchemy import or_
from typing import List, Optional
import io, csv
from ..database import get_db
from .. import models, schemas
from ..auth import get_current_user, get_visible_group_ids

router = APIRouter(prefix="/api/contacts", tags=["contacts"])


@router.get("")
def list_contacts(
    search: Optional[str] = None,
    group_id: Optional[int] = None,
    page: int = 1,
    page_size: int = 50,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    page = max(1, page)
    page_size = min(max(1, page_size), 200)

    q = db.query(models.Contact)
    gids = get_visible_group_ids(db, current_user)
    if gids is not None:
        q = q.filter(models.Contact.group_id.in_(gids))
    elif group_id:
        q = q.filter(models.Contact.group_id == group_id)
    if search:
        term = f"%{search}%"
        q = q.filter(or_(
            models.Contact.name.ilike(term),
            models.Contact.phone.ilike(term),
            models.Contact.rut_persona.ilike(term),
            models.Contact.rut_empresa.ilike(term),
            models.Contact.razon_social.ilike(term),
        ))
    q = q.order_by(models.Contact.created_at.desc())
    total = q.count()
    items = q.offset((page - 1) * page_size).limit(page_size).all()
    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
        "pages": max(1, -(-total // page_size)),  # ceil division
    }


@router.post("", response_model=schemas.ContactOut)
def create_contact(
    data: schemas.ContactCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    contact = models.Contact(**data.model_dump(), created_by=current_user.id)
    if not contact.group_id and current_user.group_id:
        contact.group_id = current_user.group_id
    db.add(contact)
    db.commit()
    db.refresh(contact)
    return contact


@router.get("/{contact_id}", response_model=schemas.ContactOut)
def get_contact(contact_id: int, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    contact = db.query(models.Contact).filter(models.Contact.id == contact_id).first()
    if not contact:
        raise HTTPException(status_code=404, detail="Contacto no encontrado")
    return contact


@router.put("/{contact_id}", response_model=schemas.ContactOut)
def update_contact(
    contact_id: int,
    data: schemas.ContactUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    contact = db.query(models.Contact).filter(models.Contact.id == contact_id).first()
    if not contact:
        raise HTTPException(status_code=404, detail="Contacto no encontrado")
    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(contact, field, value)
    db.commit()
    db.refresh(contact)
    return contact


@router.post("/bulk-import")
async def bulk_import_contacts(
    file: UploadFile = File(...),
    group_id: Optional[int] = Form(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    content = await file.read()
    filename = (file.filename or "").lower()

    rows: list[dict] = []
    errors: list[str] = []

    if filename.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
    elif filename.endswith(".xlsx") or filename.endswith(".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})
            wb.close()
        except ImportError:
            raise HTTPException(status_code=422, detail="openpyxl no instalado en el servidor")
    else:
        raise HTTPException(status_code=422, detail="Formato no soportado. Use .xlsx, .xls o .csv")

    gid = group_id or current_user.group_id
    created = 0

    FIELD_MAP = {
        "nombre": "name", "name": "name",
        "telefono": "phone", "teléfono": "phone", "phone": "phone", "celular": "phone",
        "email": "email", "correo": "email",
        "rut": "rut_persona", "rut persona": "rut_persona", "rut_persona": "rut_persona",
        "rut empresa": "rut_empresa", "rut_empresa": "rut_empresa",
        "empresa": "razon_social", "razón social": "razon_social", "razon social": "razon_social",
        "ciudad": "city", "city": "city",
        "notas": "notes", "notes": "notes", "observaciones": "notes",
    }

    for i, row in enumerate(rows, start=2):
        normalized: dict = {}
        for k, v in row.items():
            mapped = FIELD_MAP.get(k.strip().lower())
            if mapped and v:
                normalized[mapped] = v.strip()

        name = normalized.get("name", "").strip()
        phone = normalized.get("phone", "").strip()
        if not name and not phone:
            continue

        # Normalize phone
        if phone:
            phone = phone.replace(" ", "").replace("-", "").replace("(", "").replace(")", "")
            if phone.startswith("0"):
                phone = phone[1:]
            if not phone.startswith("+"):
                phone = "+56" + phone if len(phone) <= 9 else "+" + phone

        # Skip duplicates by phone
        if phone and db.query(models.Contact).filter(models.Contact.phone == phone).first():
            errors.append(f"Fila {i}: teléfono {phone} ya existe")
            continue

        try:
            contact = models.Contact(
                name=name or phone,
                phone=phone or None,
                email=normalized.get("email") or None,
                rut_persona=normalized.get("rut_persona") or None,
                rut_empresa=normalized.get("rut_empresa") or None,
                razon_social=normalized.get("razon_social") or None,
                city=normalized.get("city") or None,
                notes=normalized.get("notes") or None,
                group_id=gid,
                created_by=current_user.id,
            )
            db.add(contact)
            db.flush()
            created += 1
        except Exception as e:
            errors.append(f"Fila {i}: {str(e)}")

    db.commit()
    return {"created": created, "errors": errors}


@router.get("/export/csv")
def export_contacts_csv(
    search: Optional[str] = None,
    group_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    from fastapi.responses import StreamingResponse
    from datetime import datetime

    q = db.query(models.Contact)
    gids_csv = get_visible_group_ids(db, current_user)
    if gids_csv is not None:
        q = q.filter(models.Contact.group_id.in_(gids_csv))
    elif group_id:
        q = q.filter(models.Contact.group_id == group_id)
    if search:
        term = f"%{search}%"
        q = q.filter(or_(
            models.Contact.name.ilike(term),
            models.Contact.phone.ilike(term),
            models.Contact.rut_persona.ilike(term),
            models.Contact.rut_empresa.ilike(term),
            models.Contact.razon_social.ilike(term),
        ))
    q = q.order_by(models.Contact.created_at.desc())
    contacts = q.all()

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(["Nombre", "Teléfono", "Email", "RUT Persona", "RUT Empresa",
                     "Razón Social", "Ciudad", "Dirección", "Notas", "Creado"])
    for c in contacts:
        writer.writerow([
            c.name or "",
            c.phone or "",
            c.email or "",
            c.rut_persona or "",
            c.rut_empresa or "",
            c.razon_social or "",
            c.city or "",
            c.address or "",
            c.notes or "",
            c.created_at.strftime("%d/%m/%Y") if c.created_at else "",
        ])

    output.seek(0)
    now = datetime.now()
    filename = f"contactos_{now.strftime('%Y%m%d_%H%M%S')}.csv"
    return StreamingResponse(
        iter([output.getvalue().encode("utf-8-sig")]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.delete("/{contact_id}")
def delete_contact(
    contact_id: int,
    force: bool = False,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if current_user.role not in ("superadmin", "subadmin"):
        raise HTTPException(status_code=403, detail="Solo administradores pueden eliminar contactos")
    contact = db.query(models.Contact).filter(models.Contact.id == contact_id).first()
    if not contact:
        raise HTTPException(status_code=404, detail="Contacto no encontrado")
    active_leads = db.query(models.Lead).filter(
        models.Lead.contact_id == contact_id,
        models.Lead.current_stage.notin_(["pagado_confirmado"]),
    ).count()
    if active_leads > 0 and not (current_user.role == "superadmin" and force):
        raise HTTPException(
            status_code=400,
            detail=f"El contacto tiene {active_leads} lead(s) activo(s). Ciérrelos antes de eliminar el contacto.",
            headers={"X-Active-Leads": str(active_leads)},
        )
    db.delete(contact)
    db.commit()
    return {"ok": True}
