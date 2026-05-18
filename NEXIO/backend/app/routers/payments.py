from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from typing import List, Optional
from datetime import datetime, timezone
import io, os, uuid, asyncio, logging
from ..database import get_db
from .. import models, schemas
from ..auth import get_current_user, get_visible_group_ids
from ..utils.notifications import create_notification
from ..utils.at_informa import push_confirmed_payment, cancel_confirmed_payment

logger = logging.getLogger(__name__)

INVOICES_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../../../uploads/invoices"))
ALLOWED_TYPES = {"image/jpeg", "image/jpg", "image/png", "image/webp", "image/gif", "application/pdf"}

router = APIRouter(prefix="/api/payments", tags=["payments"])


def _payment_query(db: Session, current_user: models.User, status: Optional[str] = None, group_id: Optional[int] = None):
    q = db.query(models.PaymentVerification).options(
        joinedload(models.PaymentVerification.lead).joinedload(models.Lead.contact),
        joinedload(models.PaymentVerification.lead).joinedload(models.Lead.agendadora),
        joinedload(models.PaymentVerification.lead).joinedload(models.Lead.vendedor),
        joinedload(models.PaymentVerification.lead).joinedload(models.Lead.area),
        joinedload(models.PaymentVerification.lead).joinedload(models.Lead.group),
    )
    if current_user.role == "verificador":
        q = q.filter(models.PaymentVerification.assigned_to == current_user.id)
    elif current_user.role == "agendadora":
        q = q.join(models.Lead).filter(models.Lead.agendadora_id == current_user.id)
    elif current_user.role == "vendedor":
        q = q.join(models.Lead).filter(models.Lead.vendedor_id == current_user.id)
    elif current_user.role == "subadmin":
        gids = get_visible_group_ids(db, current_user)
        if gids is not None:
            q = q.join(models.Lead).filter(models.Lead.group_id.in_(gids))

    if status:
        q = q.filter(models.PaymentVerification.status == status)
    if group_id and current_user.role in ("superadmin", "verificador"):
        q = q.join(models.Lead, isouter=True).filter(models.Lead.group_id == group_id)

    return q


@router.get("", response_model=List[schemas.PaymentVerificationOut])
def list_payments(
    status: Optional[str] = None,
    group_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    q = _payment_query(db, current_user, status, group_id)
    return q.order_by(models.PaymentVerification.created_at.desc()).all()


@router.get("/export")
def export_payments_excel(
    status: Optional[str] = None,
    group_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """Export payments to Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    q = _payment_query(db, current_user, status, group_id)
    payments = q.order_by(models.PaymentVerification.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Verificacion de Pagos"

    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(border_style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    headers = [
        "ID", "Estado", "Cliente", "Teléfono", "RUT", "Área", "Grupo",
        "Vendedor", "Agendadora", "Honorarios", "Monto Pagado",
        "Método Pago", "Fecha Pago", "Referencia", "Notas", "Creado", "Confirmado",
    ]
    STATUS_LABELS = {"pendiente": "Pendiente", "pago_exitoso": "Confirmado", "rechazado": "Rechazado"}

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    col_widths = [5, 14, 28, 16, 14, 20, 16, 20, 20, 14, 14, 14, 12, 18, 30, 18, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    status_fill = {
        "pendiente": PatternFill("solid", fgColor="FEF3C7"),
        "pago_exitoso": PatternFill("solid", fgColor="DCFCE7"),
        "rechazado": PatternFill("solid", fgColor="FEE2E2"),
    }

    for row_idx, pv in enumerate(payments, 2):
        lead = pv.lead
        contact = lead.contact if lead else None
        row_fill = status_fill.get(pv.status)

        def cell(col, val):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.border = border
            c.alignment = Alignment(vertical="center")
            if row_fill:
                c.fill = row_fill
            return c

        cell(1, pv.id)
        cell(2, STATUS_LABELS.get(pv.status, pv.status))
        cell(3, contact.name if contact else "—")
        cell(4, contact.phone if contact else "—")
        cell(5, contact.rut_persona if contact else "—")
        cell(6, lead.area.name if lead and lead.area else "—")
        cell(7, lead.group.name if lead and lead.group else "—")
        cell(8, lead.vendedor.name if lead and lead.vendedor else "—")
        cell(9, lead.agendadora.name if lead and lead.agendadora else "—")
        cell(10, lead.honorarios if lead else 0)
        cell(11, pv.payment_amount or 0)
        cell(12, pv.payment_method or "—")
        cell(13, str(pv.payment_date) if pv.payment_date else "—")
        cell(14, pv.payment_reference or "—")
        cell(15, pv.notes or "—")
        cell(16, pv.created_at.strftime("%d/%m/%Y %H:%M") if pv.created_at else "—")
        cell(17, pv.confirmed_at.strftime("%d/%m/%Y %H:%M") if pv.confirmed_at else "—")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"pagos_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/{pv_id}/invoice")
async def upload_invoice(
    pv_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    pv = db.query(models.PaymentVerification).filter(models.PaymentVerification.id == pv_id).first()
    if not pv:
        raise HTTPException(status_code=404, detail="Verificación no encontrada")

    ct = (file.content_type or "").lower()
    if ct not in ALLOWED_TYPES:
        raise HTTPException(status_code=400, detail="Solo se permiten imágenes (JPG, PNG, WEBP) o PDF")

    os.makedirs(INVOICES_DIR, exist_ok=True)
    ext = (file.filename or "file").rsplit(".", 1)[-1].lower() if "." in (file.filename or "") else "bin"
    filename = f"pv{pv_id}_{uuid.uuid4().hex[:10]}.{ext}"
    filepath = os.path.join(INVOICES_DIR, filename)

    content = await file.read()
    if len(content) > 20 * 1024 * 1024:  # 20 MB limit
        raise HTTPException(status_code=400, detail="El archivo no puede superar 20 MB")

    with open(filepath, "wb") as f:
        f.write(content)

    pv.invoice_url = f"/uploads/invoices/{filename}"
    db.commit()
    return {"invoice_url": pv.invoice_url}


@router.put("/{pv_id}/confirm", response_model=schemas.PaymentVerificationOut)
async def confirm_payment(
    pv_id: int,
    data: schemas.PaymentVerificationUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    pv = db.query(models.PaymentVerification).options(
        joinedload(models.PaymentVerification.lead).joinedload(models.Lead.contact),
        joinedload(models.PaymentVerification.lead).joinedload(models.Lead.agendadora),
        joinedload(models.PaymentVerification.lead).joinedload(models.Lead.area),
    ).filter(models.PaymentVerification.id == pv_id).first()
    if not pv:
        raise HTTPException(status_code=404, detail="Verificación de pago no encontrada")

    if current_user.role != "verificador":
        raise HTTPException(status_code=403, detail="Solo el Verificador de Pagos puede confirmar pagos")

    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(pv, field, value)

    if data.status == "pago_exitoso":
        pv.confirmed_at = datetime.now(timezone.utc)
        lead = pv.lead
        old_stage = lead.current_stage
        lead.current_stage = "pagado_confirmado"
        db.add(models.LeadHistory(
            lead_id=lead.id,
            from_stage=old_stage,
            to_stage="pagado_confirmado",
            result="success",
            notes=f"Pago confirmado por {current_user.name}. {data.notes or ''}".strip(". "),
            created_by=current_user.id,
        ))

        for uid in {lead.agendadora_id, lead.vendedor_id} - {None}:
            create_notification(
                db, uid,
                "Pago confirmado ✓",
                f"El pago de {lead.contact.name} fue CONFIRMADO por {current_user.name}.",
                lead_id=lead.id,
                notification_type="pago_confirmado"
            )

        contact = lead.contact
        if contact and contact.phone:
            db.add(models.WhatsAppMessage(
                lead_id=lead.id,
                contact_id=contact.id,
                direction="out",
                message_type="text",
                content=f"Estimado/a {contact.name}, le informamos que su pago ha sido confirmado exitosamente. Gracias por su confianza en nuestros servicios.",
                status="pending",
                sent_by=current_user.id,
            ))

    elif data.status == "rechazado":
        lead = pv.lead
        old_stage = lead.current_stage
        lead.current_stage = "recuperacion_cierre"
        db.add(models.LeadHistory(
            lead_id=lead.id,
            from_stage=old_stage,
            to_stage="recuperacion_cierre",
            result="failed",
            notes=f"Pago rechazado por {current_user.name}. {data.notes or ''}".strip(". "),
            created_by=current_user.id,
        ))

        for uid in {lead.agendadora_id, lead.vendedor_id} - {None}:
            create_notification(
                db, uid,
                "Pago rechazado",
                f"El pago de {lead.contact.name} fue RECHAZADO. El lead vuelve a Recuperación Cierre.",
                lead_id=lead.id,
                notification_type="pago"
            )

    # Capture data needed for AT Informa before commit flushes the session
    _at_payload = None
    if data.status == "pago_exitoso":
        lead = pv.lead
        contact = lead.contact
        _at_payload = {
            "full_name":   contact.name,
            "email":       contact.email or f"{contact.phone}@sin-email.crm",
            "phone":       contact.phone or "",
            "category":    lead.area.name if lead.area else "General",
            "invoice_url": pv.invoice_url or None,
            "case_code":   f"CRM-{lead.id:05d}",
        }

    db.commit()
    db.refresh(pv)

    # Fire AT Informa integration (non-blocking — don't fail the CRM op if AT is down)
    if _at_payload:
        try:
            result = await push_confirmed_payment(**_at_payload)
            logger.info("[AT Informa] Caso creado: %s", result)
        except Exception as exc:
            logger.warning("[AT Informa] Integración falló (el pago CRM sigue confirmado): %s", exc)

    return db.query(models.PaymentVerification).options(
        joinedload(models.PaymentVerification.lead).joinedload(models.Lead.contact),
        joinedload(models.PaymentVerification.lead).joinedload(models.Lead.agendadora),
        joinedload(models.PaymentVerification.lead).joinedload(models.Lead.vendedor),
        joinedload(models.PaymentVerification.lead).joinedload(models.Lead.area),
        joinedload(models.PaymentVerification.lead).joinedload(models.Lead.group),
    ).filter(models.PaymentVerification.id == pv_id).first()


@router.post("/{pv_id}/revert", response_model=schemas.PaymentVerificationOut)
async def revert_payment(
    pv_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if current_user.role != "verificador":
        raise HTTPException(status_code=403, detail="Solo el Verificador de Pagos puede revertir pagos")

    pv = db.query(models.PaymentVerification).options(
        joinedload(models.PaymentVerification.lead).joinedload(models.Lead.contact),
        joinedload(models.PaymentVerification.lead).joinedload(models.Lead.agendadora),
        joinedload(models.PaymentVerification.lead).joinedload(models.Lead.vendedor),
    ).filter(models.PaymentVerification.id == pv_id).first()

    if not pv:
        raise HTTPException(status_code=404, detail="Verificación de pago no encontrada")
    if pv.status != "pago_exitoso":
        raise HTTPException(status_code=400, detail="Solo se pueden revertir pagos confirmados")

    # Find previous lead stage from history
    prev_entry = (
        db.query(models.LeadHistory)
        .filter(models.LeadHistory.lead_id == pv.lead_id, models.LeadHistory.to_stage == "pagado_confirmado")
        .order_by(models.LeadHistory.id.desc())
        .first()
    )
    revert_stage = prev_entry.from_stage if prev_entry else "cierre"

    # Revert payment
    pv.status = "pendiente"
    pv.confirmed_at = None
    pv.payment_amount = None
    pv.payment_method = None
    pv.payment_date = None
    pv.payment_reference = None
    pv.notes = f"[REVERTIDO por {current_user.name}] " + (pv.notes or "")

    # Revert lead stage
    lead = pv.lead
    old_stage = lead.current_stage
    lead.current_stage = revert_stage
    db.add(models.LeadHistory(
        lead_id=lead.id,
        from_stage=old_stage,
        to_stage=revert_stage,
        result="reverted",
        notes=f"Pago revertido a pendiente por {current_user.name}. Vuelve a etapa anterior.",
        created_by=current_user.id,
    ))

    contact_name = lead.contact.name if lead.contact else "el cliente"
    for uid in {lead.agendadora_id, lead.vendedor_id} - {None}:
        create_notification(
            db, uid,
            "Pago revertido ↩",
            f"El pago confirmado de {contact_name} fue REVERTIDO a pendiente por {current_user.name}.",
            lead_id=lead.id,
            notification_type="pago"
        )

    _at_case_code = f"CRM-{pv.lead_id:05d}"
    db.commit()

    # Notify AT Informa to revert the case (non-blocking)
    try:
        result = await cancel_confirmed_payment(case_code=_at_case_code)
        logger.info("[AT Informa] Caso revertido: %s", result)
    except Exception as exc:
        logger.warning("[AT Informa] Cancelación falló (el pago CRM sigue revertido): %s", exc)

    return db.query(models.PaymentVerification).options(
        joinedload(models.PaymentVerification.lead).joinedload(models.Lead.contact),
        joinedload(models.PaymentVerification.lead).joinedload(models.Lead.agendadora),
        joinedload(models.PaymentVerification.lead).joinedload(models.Lead.vendedor),
        joinedload(models.PaymentVerification.lead).joinedload(models.Lead.area),
        joinedload(models.PaymentVerification.lead).joinedload(models.Lead.group),
    ).filter(models.PaymentVerification.id == pv_id).first()
